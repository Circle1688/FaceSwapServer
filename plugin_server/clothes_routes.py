import json

from fastapi import APIRouter, Depends, HTTPException, status

from plugin_server.models import *
from plugin_server.database_func import get_db
from sqlalchemy.orm import Session
from fastapi import UploadFile, File
from plugin_server.schemas import *

from openpyxl import load_workbook
from io import BytesIO
from sqlalchemy import or_

router = APIRouter()

@router.post('/upload_clothes')
async def upload_clothes(file: UploadFile = File(...), db: Session = Depends(get_db)):
	# 检查文件类型
	if not file.filename.endswith(('.xls', '.xlsx')):
		raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="The file format is incorrect. Only Excel files are supported")

	# 清空数据库表
	db.query(Clothes).delete()
	db.commit()

	# 将 SpooledTemporaryFile 对象的内容读取到 BytesIO 对象中
	contents = file.file.read()
	file.file.close()
	excel_file = BytesIO(contents)

	# 读取 Excel 文件内容
	wb = load_workbook(excel_file)
	ws = wb.active

	# 解析 Excel 并写入数据库
	for row in ws.iter_rows(min_row=2, values_only=True):  # 假设第一行是表头
		try:
			url, brand, gender, name, colors, colors_hex, sizes = row
			# 将字符串转换为数组
			colors = json.dumps(colors.split(",") if isinstance(colors, str) else [])
			colors_hex = json.dumps(colors_hex.split(",") if isinstance(colors_hex, str) else [])
			sizes = json.dumps(sizes.split(",") if isinstance(sizes, str) else [])

			clothes = Clothes(url=url, brand=brand, gender=gender, name=name, colors=colors, colors_hex=colors_hex, sizes=sizes)
			db.add(clothes)
		except Exception as e:
			raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Error processing row {row}: {str(e)}")

	db.commit()

	return {"message": "The Excel file has been uploaded and parsed successfully, and data has been written to the database"}

@router.post('/get_clothes')
async def get_clothes(request: ClothesRequest, db: Session = Depends(get_db)):
	db_clothes = db.query(Clothes).filter(or_(Clothes.url == request.url, (Clothes.brand == request.brand) & (Clothes.name == request.name))).first()
	if not db_clothes:
		raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No data found")

	# 将数组字段转换为字符串
	db_clothes.colors = json.loads(db_clothes.colors)
	db_clothes.colors_hex = json.loads(db_clothes.colors_hex)
	db_clothes.sizes = json.loads(db_clothes.sizes)

	result = db_clothes.__dict__
	result.pop('url')
	result.pop('id')

	return result

